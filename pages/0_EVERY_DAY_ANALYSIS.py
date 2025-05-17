#  orginal file name is 2_FILE_EXPORT_NEW_DESIGN_YAHOO.py and changed to 0_FILE_EXPORT_TO_STOCK_SEARCH
import streamlit as st
import pandas as pd
import yfinance as yf
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
from io import BytesIO

def create_excel_with_hyperlinks(df):
    output = BytesIO()  # Create a BytesIO object to hold the Excel file
    workbook = openpyxl.Workbook()  # Create a new Excel workbook
    worksheet = workbook.active  # Get the active worksheet

    # Write the header
    for col_num, column_title in enumerate(df.columns, 1):
        worksheet.cell(row=1, column=col_num, value=column_title)

    # Write the data with hyperlinks
    for row_num, row in enumerate(df.itertuples(index=False), 2):
        for col_num, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            if df.columns[col_num - 1] == 'Last Sale':
                # Create the hyperlink for the last sale price
                symbol = row.Symbol  # Get the symbol directly from the row
                hyperlink = f'https://finance.yahoo.com/quote/{symbol}/'
                cell.value = value  # Set the display text to the last sale price (e.g., 763.89)
                cell.hyperlink = hyperlink  # Set the hyperlink without any HTML tags
                cell.style = 'Hyperlink'  # Optional: style it as a hyperlink
            else:
                cell.value = value  # For other columns, just set the value

    workbook.save(output)  # Save the workbook to the BytesIO object
    output.seek(0)  # Move the cursor to the beginning of the BytesIO object
    return output.getvalue()  # Return the Excel file content


# Initialize session state variables if not already set
if 'filter_data' not in st.session_state:
    st.session_state.filter_data = None
if 'selected_symbol' not in st.session_state:
    st.session_state.selected_symbol = None
if 'current_page' not in st.session_state:
    st.session_state.current_page = 1
if 'time_frame' not in st.session_state:
    st.session_state.time_frame = '5D'  # Default time frame

def get_historical_data(symbol_input, period):
    """
    Fetch historical stock data for the given symbol and period.
    :param symbol_input: Stock ticker symbol.
    :param period: Time period for the data.
    :return: DataFrame containing historical data.
    """
    ticker = yf.Ticker(symbol_input)
    history = ticker.history(period=period)
    return history

def plot_stock_prices(history, symbol):
    """
    Plot historical stock prices using Plotly with technical indicators and time frame selection.
    :param history: DataFrame containing historical stock data.
    :param symbol: Stock symbol for plotting.
    """
    # Convert the index to datetime if it's not already
    if not isinstance(history.index, pd.DatetimeIndex):
        history.index = pd.to_datetime(history.index)
    
    fig = make_subplots(rows=2, cols=1, shared_xaxes=True,
                        vertical_spacing=0.1,
                        subplot_titles=(f'{symbol} Stock Price', 'Technical Indicators'),
                        row_heights=[0.7, 0.3])
    
    # Add OHLC (Open, High, Low, Close) traces
    fig.add_trace(go.Candlestick(x=history.index.strftime('%Y-%m-%d'),
                                 open=history['Open'],
                                 high=history['High'],
                                 low=history['Low'],
                                 close=history['Close'],
                                 name='OHLC'),
                  row=1, col=1)

    # Add moving averages
    for ma in [10, 50, 200]:
        if ma <= len(history):
            history[f'MA_{ma}'] = history['Close'].rolling(window=ma).mean()
            fig.add_trace(go.Scatter(x=history.index.strftime('%Y-%m-%d'), y=history[f'MA_{ma}'], mode='lines', name=f'MA {ma}'),
                          row=1, col=1)

    # Add RSI trace
    delta = history['Close'].diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
    rs = gain / loss
    rsi = 100 - (100 / (1 + rs))
    fig.add_trace(go.Scatter(x=history.index.strftime('%Y-%m-%d'), y=rsi, mode='lines', name='RSI'),
                  row=2, col=1)

    # Add RSI markers
    overbought = 70
    oversold = 30
    fig.add_trace(go.Scatter(x=history.index[rsi > overbought].strftime('%Y-%m-%d'), y=rsi[rsi > overbought],
                             mode='markers', marker=dict(symbol='triangle-up', color='red'), name='Overbought'),
                  row=2, col=1)
    fig.add_trace(go.Scatter(x=history.index[rsi < oversold].strftime('%Y-%m-%d'), y=rsi[rsi < oversold],
                             mode='markers', marker=dict(symbol='triangle-down', color='green'), name='Oversold'),
                  row=2, col=1)

    # Add additional technical indicators: MACD and Bollinger Bands
    # MACD
    short_ema = history['Close'].ewm(span=12, adjust=False).mean()
    long_ema = history['Close'].ewm(span=26, adjust=False).mean()
    macd = short_ema - long_ema
    macd_signal = macd.ewm(span=9, adjust=False).mean()
    fig.add_trace(go.Scatter(x=history.index.strftime('%Y-%m-%d'), y=macd, mode='lines', name='MACD'),
                  row=2, col=1)
    fig.add_trace(go.Scatter(x=history.index.strftime('%Y-%m-%d'), y=macd_signal, mode='lines', name='MACD Signal'),
                  row=2, col=1)

    # Bollinger Bands
    window = 20
    std_dev = history['Close'].rolling(window=window).std()
    bollinger_upper = history['Close'].rolling(window=window).mean() + (std_dev * 2)
    bollinger_lower = history['Close'].rolling(window=window).mean() - (std_dev * 2)
    fig.add_trace(go.Scatter(x=history.index.strftime('%Y-%m-%d'), y=bollinger_upper, mode='lines', name='Bollinger Upper'),
                  row=1, col=1)
    fig.add_trace(go.Scatter(x=history.index.strftime('%Y-%m-%d'), y=bollinger_lower, mode='lines', name='Bollinger Lower'),
                  row=1, col=1)

    fig.update_layout(title=f'Historical Stock Prices for {symbol}',
                      xaxis_title='Date',
                      yaxis_title='Price',
                      hovermode='x unified',
                      xaxis_rangeslider_visible=False)

    # Assuming st is your Streamlit import
    st.plotly_chart(fig)


def get_yahoo_finance_link(symbol):
    return f'https://finance.yahoo.com/quote/{symbol}/'

def get_stock_data(df, min_amount, max_amount, profit_loss, min_volume, max_volume, last_sale_min=None, last_sale_max=None):
    """
    Filter stock data based on profit/loss criteria, volume, and last sale price, and convert results to DataFrame.
    :param df: DataFrame containing stock data.
    :param min_amount: Minimum amount for profit/loss filtering.
    :param max_amount: Maximum amount for profit/loss filtering.
    :param profit_loss: 'profit' or 'loss' to filter by.
    :param min_volume: Minimum volume for filtering.
    :param max_volume: Maximum volume for filtering.
    :param last_sale_min: Minimum last sale price for filtering.
    :param last_sale_max: Maximum last sale price for filtering.
    :return: DataFrame of filtered and processed stock data.
    """
    results = []

    df['last sale'] = df['last sale'].replace(r'\$', '', regex=True).astype(float)
    df['net change'] = pd.to_numeric(df['net change'], errors='coerce')
    df['volume'] = pd.to_numeric(df['volume'], errors='coerce')
    
    df['% change'] = df.get('% change', pd.Series()).str.replace('%', '').astype(float)
    df['market cap'] = df.get('market cap', pd.Series())
    df['ipo year'] = df.get('ipo year', pd.Series())
    df['sector'] = df.get('sector', pd.Series())
    df['industry'] = df.get('industry', pd.Series())
    df['country'] = df.get('country', pd.Series())

    df = df.dropna(subset=['symbol', 'last sale', 'net change'])

    if min_volume is not None and max_volume is not None:
        df = df[(df['volume'] >= min_volume) & (df['volume'] <= max_volume)]
    
    if last_sale_min is not None:
        df = df[df['last sale'] >= last_sale_min]
    
    if last_sale_max is not None:
        df = df[df['last sale'] <= last_sale_max]

    total_symbols = len(df)
    progress_bar = st.progress(0)

    if total_symbols > 0:
        for i, row in df.iterrows():
            try:
                symbol = row['symbol']
                last_sale = row['last sale']
                net_change = row['net change']
                percent_change = row['% change'] if pd.notna(row['% change']) else 0.0
                market_cap = row['market cap'] if pd.notna(row['market cap']) else "N/A"
                ipo_year = int(row['ipo year']) if pd.notna(row['ipo year']) else "N/A"
                volume = row['volume'] if pd.notna(row['volume']) else "N/A"
                sector = row['sector'] if pd.notna(row['sector']) else "N/A"
                industry = row['industry'] if pd.notna(row['industry']) else "N/A"
                country = row['country'] if pd.notna(row['country']) else "N/A"

                if (profit_loss == 'profit' and min_amount <= net_change <= max_amount) or \
                   (profit_loss == 'loss' and -max_amount <= net_change <= -min_amount):
                    yahoo_finance_link = get_yahoo_finance_link(symbol)
                    
                    results.append({
                        "Symbol": symbol,
                        "Last Sale": f'<a href="{yahoo_finance_link}" target="_blank">{last_sale:.2f}</a>',
                        "Net Change": net_change,
                        "Percent Change": percent_change,
                        "Market Cap": market_cap,
                        "IPO Year": ipo_year,
                        "Volume": volume,
                        "Sector": sector,
                        "Industry": industry,
                        "Country": country
                    })
            except Exception as e:
                st.error(f"Error processing symbol {symbol}: {e}")

            progress_value = (i + 1) / total_symbols
            progress_bar.progress(min(max(progress_value, 0.0), 1.0))

    results_df = pd.DataFrame(results)

    if results_df.empty:
        st.write("No stocks meet the specified criteria.")
        return pd.DataFrame()

    results_df['Net Change'] = pd.to_numeric(results_df['Net Change'], errors='coerce')
    results_df = results_df.sort_values(by='Net Change', ascending=False)
    
    return results_df

def get_advisory(history):
    """
    Generate stock advisory based on technical indicators.
    :param history: DataFrame containing historical stock data.
    :return: Advisory string.
    """
    advisory = "No advisory available."

    if history.empty or len(history) < 30:  # Ensure there's enough data to compute indicators
        return "Insufficient data for advisory."

    # Calculate RSI
    delta = history['Close'].diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
    rs = gain / loss
    rsi = 100 - (100 / (1 + rs))

    if not rsi.empty:
        latest_rsi = rsi.iloc[-1]
    else:
        latest_rsi = None

    # Calculate MACD
    short_ema = history['Close'].ewm(span=12, adjust=False).mean()
    long_ema = history['Close'].ewm(span=26, adjust=False).mean()
    macd = short_ema - long_ema
    macd_signal = macd.ewm(span=9, adjust=False).mean()

    if not macd.empty and not macd_signal.empty:
        latest_macd = macd.iloc[-1]
        latest_macd_signal = macd_signal.iloc[-1]
    else:
        latest_macd = None
        latest_macd_signal = None

    # Calculate Bollinger Bands
    window = 20
    std_dev = history['Close'].rolling(window=window).std()
    bollinger_upper = history['Close'].rolling(window=window).mean() + (std_dev * 2)
    bollinger_lower = history['Close'].rolling(window=window).mean() - (std_dev * 2)

    if not bollinger_upper.empty and not bollinger_lower.empty:
        latest_close = history['Close'].iloc[-1]
        latest_bollinger_upper = bollinger_upper.iloc[-1]
        latest_bollinger_lower = bollinger_lower.iloc[-1]
    else:
        latest_close = None
        latest_bollinger_upper = None
        latest_bollinger_lower = None

    # Advisory based on RSI
    if latest_rsi is not None:
        if latest_rsi > 70:
            advisory = "Sell. The stock is in an overbought condition."
        elif latest_rsi < 30:
            advisory = "Buy. The stock is in an oversold condition."

    # Advisory based on MACD
    if latest_macd is not None and latest_macd_signal is not None:
        if latest_macd < latest_macd_signal:
            advisory = "Sell. The MACD line is below the signal line."
        elif latest_macd > latest_macd_signal:
            advisory = "Buy. The MACD line is above the signal line."

    # Advisory based on Bollinger Bands
    if latest_close is not None and latest_bollinger_upper is not None and latest_bollinger_lower is not None:
        if latest_close > latest_bollinger_upper:
            advisory = "Sell. The stock price is above the upper Bollinger Band."
        elif latest_close < latest_bollinger_lower:
            advisory = "Buy. The stock price is below the lower Bollinger Band."

    return advisory


def main():
    st.set_page_config(layout="wide")

    st.title("Stock Price Analysis")
    st.header("Excel Data Filtering")
    st.markdown("""
    **Note:** Use this [link](https://www.nasdaq.com/market-activity/stocks/screener) to get the .csv file for upload and analysis.
    """)

    st.sidebar.header("Upload and Filter Data")
    uploaded_file = st.sidebar.file_uploader("Upload file", type=["csv"])

    if uploaded_file is not None:
        df = pd.read_csv(uploaded_file)
        df.columns = df.columns.str.lower()

        column_mapping = {
            'symbol': 'symbol',
            'last sale': 'last sale',
            'net change': 'net change',
            '% change': '% change',
            'market cap': 'market cap',
            'ipo year': 'ipo year',
            'volume': 'volume',
            'sector': 'sector',
            'industry': 'industry',
            'country': 'country'
        }
        df.rename(columns=column_mapping, inplace=True)

        required_columns = ['symbol', 'last sale', 'net change']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            st.sidebar.error(f"Uploaded file must contain the following columns: {', '.join(required_columns)}")
            return

        filter_sector = st.sidebar.selectbox("Filter by sector (optional):", options=[""] + df.get('sector', pd.Series()).dropna().unique().tolist())
        #filter_industry = st.sidebar.selectbox("Filter by industry (optional):", options=[""] + df.get('industry', pd.Series()).dropna().unique().tolist())
        industries = df[df['sector'].str.lower() == filter_sector.lower()]['industry'].dropna().unique().tolist() if filter_sector else []
        filter_industry = st.sidebar.selectbox("Filter by industry (optional):", options=[""] + industries)
        profit_loss = st.sidebar.selectbox("Select profit/loss:", ['profit', 'loss'])
        min_amount = st.sidebar.number_input(f"Enter minimum {profit_loss} amount:", format="%.2f", value=0.0)
        max_amount = st.sidebar.number_input(f"Enter maximum {profit_loss} amount:", format="%.2f", value=1000.0)

        volume_options = df['volume'].dropna().unique()
        volume_options.sort()
        min_volume = st.sidebar.selectbox("Select minimum volume:", options=[None] + list(volume_options))
        max_volume = st.sidebar.selectbox("Select maximum volume:", options=[None] + list(volume_options))
            
        last_sale_min = st.sidebar.number_input("Last Sale Min:", format="%.2f", value=0.0)
        last_sale_max = st.sidebar.number_input("Last Sale Max:", format="%.2f", value=1000.0)

        if st.sidebar.button("Apply Filter"):
            if filter_sector:
                df = df[df['sector'].str.lower() == filter_sector.lower()]
            if filter_industry:
                df = df[df['industry'].str.lower() == filter_industry.lower()]

            if df.empty:
                st.write("No data available for the selected filters.")
                return

            results_df = get_stock_data(df, min_amount, max_amount, profit_loss, min_volume, max_volume, last_sale_min, last_sale_max)
            
            if not results_df.empty:
                st.session_state.filter_data = results_df
                st.session_state.current_page = 1  # Reset to first page on new filter application
                
                st.write(f"**{len(results_df)} records fetched.**")
                # Add download button for Excel
                excel_data = create_excel_with_hyperlinks(results_df)
                st.download_button(
                    label="Download Filtered Data",
                    data=excel_data,
                    file_name='filtered_stock_data.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
            else:
                st.session_state.filter_data = None

    # Time frame selection with expander
    with st.sidebar.expander("Select Time Frame"):
        time_frame = st.radio("Select Time Frame:", [ "5D", "1MO", "6MO", "YTD", "1Y","2Y","5Y", "max"])
        st.session_state.time_frame = time_frame  # Save selected time frame

    filter_data = st.session_state.get('filter_data', None)
    if filter_data is not None:
        # Pagination
        page_size = 30  # Set page size to 30 as per your requirement
        num_pages = (len(filter_data) + page_size - 1) // page_size
        st.session_state.current_page = st.selectbox('Select page:', range(1, num_pages + 1), index=st.session_state.current_page-1)
        
        # Display table for selected page
        start_idx = (st.session_state.current_page - 1) * page_size
        end_idx = min(start_idx + page_size, len(filter_data))
        
        st.write(filter_data.iloc[start_idx:end_idx].to_html(escape=False, index=False,
                    formatters={'Last Sale': lambda x: f'<a href="{get_yahoo_finance_link(filter_data.iloc[start_idx]["Symbol"])}" target="_blank">{x}</a>'},
                    classes='scrollable-table'), unsafe_allow_html=True)

        selected_symbol = st.selectbox("Select a stock to view details:", options=filter_data['Symbol'].tolist(), index=0)
        st.session_state.selected_symbol = selected_symbol

        if st.session_state.selected_symbol:
            st.write(f"Fetching historical data for {st.session_state.selected_symbol}...")
            history = get_historical_data(st.session_state.selected_symbol, st.session_state.time_frame)
            plot_stock_prices(history, st.session_state.selected_symbol)
            st.write(f"Historical Data for {st.session_state.selected_symbol}")
            st.dataframe(history)

            # Advisory based on plot data
            advisory = get_advisory(history)
            st.write(advisory)


            #REFER STOCK DATA FOR ANALYSIS
            st.write("Refer below program for stock Data Analysis")
            st.markdown("""
            - [Stock Data Analysis](https://doondestockanalysis-mpqqyujrvtmdcqdx253syu.streamlit.app/)
             """)
            
            # Collapsible section for referral links
            with st.expander("Referral Links for Further Analysis"):
                st.markdown("""
                - [Finviz](https://finviz.com/)
                - [Financhill](https://financhill.com/)
                - [Ainvest Screener](https://www.ainvest.com/screener/)
                - [Investing.com](https://www.investing.com/)
                - [CNBC US Markets](https://www.cnbc.com/us-markets/)
                - [StockFetcher Chart](https://www.stockfetcher.com/html5_chart/chartplus.php)
                - [Fedility](https://www.fidelity.com/)
                - [Investopedia](https://www.investopedia.com/trading-4427765)
                - [Sec.gov](https://www.sec.gov/search-filings/)
                """)
            

    if st.sidebar.button("Clear Filters"):
        st.session_state.filter_data = None
        st.session_state.selected_symbol = None
        st.session_state.current_page = 1
        st.session_state.time_frame = 'YTD'  # Reset to default time frame
        st.rerun()

   
if __name__ == '__main__':
    main()

